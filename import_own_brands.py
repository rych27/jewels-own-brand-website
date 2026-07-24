import json
from openpyxl import load_workbook

INPUT = '/Users/mjadrych/Downloads/OB is Highlighted Green (1).xlsx'
OUTPUT = 'own-brand-products.js'

# Order matters: more specific / easily-confused categories are checked
# first so generic keywords in later categories (e.g. "chicken" in "chicken
# gravy", "cat" in "cat food") can't steal items that belong elsewhere.
CATEGORY_RULES = [
    ('Pet Care', [
        'CAT FOOD', 'CAT LITTER', 'CAT CHOW', 'CAT TREAT', 'CAT ',
        'DOG FOOD', 'DOG TREAT', 'DOG ', 'PET ', 'KITTEN', 'PUPPY',
        'FELINE', 'CANINE', 'PURINA', 'FRISKIES', 'PEDIGREE', 'IAMS',
        'BLUE BUFFALO', 'FANCY FEAST', 'MEOW MIX', 'KIBBLE', 'CAT LITTER BOX'
    ]),
    ('Household', [
        'DETERGENT', 'CLEANER', 'PAPER TOWEL', 'TOILET PAPER',
        'BATHROOM TISSUE', 'NAPKIN', 'TRASH BAG', 'ALUMINUM FOIL',
        'PLASTIC WRAP', 'FOIL', 'BAGGIES', 'ZIPLOC', 'DISINFECT',
        'BLEACH', 'HAND SOAP', 'BODY SOAP', 'BAR SOAP', 'SPONGE',
        'LAUNDRY', 'FABRIC SOFTENER', 'DISH SOAP', 'DISHWASHER',
        'AIR FRESHENER', 'CANDLE', 'BATTERIES', 'LIGHT BULB'
    ]),
    ('Sauces & Condiments', [
        'GRAVY', 'BROTH', 'STOCK', 'PASTA SAUCE', 'BBQ SAUCE',
        'HOT SAUCE', 'SOY SAUCE', 'SALSA', 'SAUCE', 'MARINADE',
        'SEASONING', 'SPICE', 'BOUILLON', 'KETCHUP', 'MUSTARD',
        'MAYO', 'MAYONNAISE', 'SALAD DRESSING', 'VINAIGRETTE', 'RELISH'
    ]),
    ('Dairy', [
        'MILK', 'CHEESE', 'YOGURT', 'YOGHURT', 'EGG', 'BUTTER',
        'CREAMER', 'SOUR CREAM', 'COTTAGE CHEESE', 'CREAM CHEESE',
        'WHIPPED CREAM', 'HALF AND HALF'
    ]),
    ('Bakery', [
        'BREAD', 'BAGEL', 'DINNER ROLL', 'ROLL', 'CAKE', 'COOKIE',
        'MUFFIN', 'PIE', 'DONUT', 'DOUGHNUT', 'TORTILLA', 'CROISSANT',
        'PASTRY', 'BISCUIT'
    ]),
    ('Deli', [
        'DELI', 'ROTISSERIE', 'PREPARED FOOD', 'LUNCH MEAT',
        'SLICED HAM', 'SLICED TURKEY', 'CHARCUTERIE'
    ]),
    ('Frozen', [
        'FROZEN', 'ICE CREAM', 'FREEZER', 'POPSICLE', 'ICE POP'
    ]),
    ('Meat & Seafood', [
        'GROUND BEEF', 'STEAK', 'BEEF', 'CHICKEN BREAST', 'CHICKEN THIGH',
        'CHICKEN WING', 'WHOLE CHICKEN', 'CHICKEN', 'PORK CHOP', 'BACON',
        'SAUSAGE', 'PORK', 'TURKEY', 'LAMB', 'SEAFOOD', 'SALMON', 'TUNA',
        'FISH', 'SHRIMP', 'CRAB', 'LOBSTER'
    ]),
    ('Beverages', [
        'BOTTLED WATER', 'SPARKLING WATER', 'WATER', 'SODA', 'POP',
        'FRUIT JUICE', 'JUICE', 'COFFEE', 'TEA', 'ENERGY DRINK',
        'SPORTS DRINK', 'BEER', 'WINE'
    ]),
    ('Produce', [
        'APPLE', 'BANANA', 'BERRY', 'STRAWBERRY', 'BLUEBERRY',
        'AVOCADO', 'SALAD', 'LETTUCE', 'ONION', 'POTATO', 'TOMATO',
        'CARROT', 'PEPPER', 'CUCUMBER', 'VEGETABLE', 'FRUIT'
    ]),
]


def category(link, description):
    text = f'{link} {description}'.upper()
    for cat_name, keywords in CATEGORY_RULES:
        if any(kw in text for kw in keywords):
            return cat_name
    return 'Pantry'


def emoji(cat):
    return {
        'Dairy': '🥛',
        'Bakery': '🥖',
        'Deli': '🥪',
        'Frozen': '🧊',
        'Meat & Seafood': '🥩',
        'Beverages': '🥤',
        'Produce': '🍓',
        'Pantry': '🥫',
        'Sauces & Condiments': '🍯',
        'Pet Care': '🐾',
        'Household': '🧻'
    }.get(cat, '🥫')


ws = load_workbook(INPUT, data_only=True).active
products = []

for row in ws.iter_rows(min_row=2, values_only=False):
    desc_cell = row[1]
    link, description, _, price = [cell.value for cell in row[:4]]
    if not description:
        continue

    item = str(description).strip()
    own_brand = bool(desc_cell.fill and desc_cell.fill.start_color and desc_cell.fill.start_color.rgb in ['FF00FF00', '00FF00', '0000FF00', 'FF00FF00'])
    cat = category(str(link or ''), item)
    name = item.title()
    usable_price = float(price) if isinstance(price, (int, float)) and price > 0 else None

    products.append({
        'id': len(products) + 10001,
        'name': name,
        'brand': 'Own Brands' if own_brand else 'National Brand',
        'brandType': 'own' if own_brand else 'national',
        'price': usable_price,
        'cat': cat,
        'sourceCategory': str(link or ''),
        'emoji': emoji(cat),
        'tag': 'OWN BRAND' if own_brand else ''
    })

# Compute comparison prices
national_prices_by_group = {}
for p in products:
    if p['brandType'] == 'national' and p['price'] is not None:
        national_prices_by_group.setdefault(p['sourceCategory'], []).append(p['price'])

for p in products:
    if p['brandType'] == 'own' and p['price'] is not None:
        nat_prices = national_prices_by_group.get(p['sourceCategory'])
        if nat_prices:
            avg_nat = round(sum(nat_prices) / len(nat_prices), 2)
            if avg_nat > p['price']:
                p['natPrice'] = avg_nat

with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write('window.OWN_BRAND_PRODUCTS = ')
    json.dump(products, f, ensure_ascii=False, indent=2)

print("Export complete!")
